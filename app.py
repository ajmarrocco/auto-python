import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    # Loads excel workbook
    wb = xl.load_workbook('transactions.xlsx')
    # Gets sheet 
    sheet = wb['Sheet1']
    # Gets cell
    cell = sheet['a1']

    for row in range(2, sheet.max_row + 1):
        # gets what is in each cell in the row
        # .cell(row, column)
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row,4)
        corrected_price_cell.value = corrected_price


    values = Reference(sheet, 
                        min_row =2, 
                        max_row = sheet.max_row,
                        min_col=4,
                        max_col = 4
                        )

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    # Saves in new file 
    wb.save('transactions.xlsx')
